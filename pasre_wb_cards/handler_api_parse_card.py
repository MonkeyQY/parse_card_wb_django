import asyncio
import json
import logging
from typing import Any

import aiohttp
from openpyxl.reader.excel import load_workbook
from openpyxl.workbook import Workbook
from rest_framework.response import Response

from pasre_wb_cards.sku_model import SkuModel

log = logging.getLogger("Handling Parse WB Cards")


async def process_numbers(request):
    number = request.data["sku"]
    # обработка данных

    async with aiohttp.ClientSession() as session:
        sku = await get_sku_info(session, number)

    log.info(f"Полученные данные для {number}: {sku}")

    return Response({"status": "success", "result": sku})


def process_file(request):
    try:
        uploaded_file = request.FILES["file"]
        list_sku = read_excel(uploaded_file)
    except Exception as e:
        log.error(f"Ошибка при загрузке файла: {e}")
        return Response({"status": "error", "message": "File not read"})

    dict_sku = asyncio.run(parse_sku(list_sku))

    log.info(f"Полученные данные: {dict_sku}")
    return Response({"status": "success", "result": dict_sku})


def read_excel(excel_file: Any) -> list:
    wb: Workbook = load_workbook(excel_file)
    ws = wb.active

    list_sku = [x for x in ws["A"]]

    return list_sku


async def parse_sku(list_sku: list) -> dict:
    # TODO : ОБЯЗТЕЛЬНО НУЖЕН ПРОКСИ, ЕСЛИ ЗАГРУЖАЕТСЯ МНОГО СКУ

    dict_sku = {}
    async with aiohttp.ClientSession() as session:
        async with asyncio.TaskGroup() as tg:
            for sku in list_sku:
                dict_sku[sku] = tg.create_task(get_sku_info(session, sku))
    return dict_sku


async def get_sku_info(session: aiohttp.ClientSession, sku: str) -> SkuModel:
    url = (
        f"https://card.wb.ru/cards/detail?appType=128&curr=rub&locale="
        f"by&lang=ru&dest=-59208&"
        f"regions=1,4,22,30,31,33,40,48,66,68,69,70,80,83&reg=1&spp=0&nm={sku}"
    )
    async with session.get(url) as response:
        data = json.loads(await response.text())
        sku = {
            "id": data["data"]["products"][0]["id"],
            "name": data["data"]["products"][0]["name"],
            "brand": data["data"]["products"][0]["brand"],
        }
        return SkuModel.parse_obj(sku)
